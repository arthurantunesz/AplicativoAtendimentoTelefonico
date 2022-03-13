from tkinter import *
from datetime import datetime
import tkinter as tk
from tkinter import Tk
from tkinter import ttk
import pandas as pd
from PIL import ImageTk, Image
import tkinter.font as tkFont
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pyautogui

lista_usuarios = {'arthur':['Arthur','a797'],
                  'larissa':['Larissa','l561'],
                  'tamiris':['Tamiris','t494'],
                  'jessica':['Jéssica','j162']}

lista_motivo = ['1. Vender o consorcio/conhecer como trabalhamos e calculamos o preço','2. Engano. Queria ligar para o banco/outra adm','3. Retorno sobre o preenchimento do formulário',
                    '4. Status da cota/prazo de pagamento','5. Problemas técnicos com plataforma fastin ou dados incorretos','6. Cliente quer começar/comprar um consórcio','7. Outro']

lista_adm = ['Bradesco', 'Caixa', 'Canopus', 'Itaú', 'Magalu/Luiza', 'Porto Seguro', 'Randon', 'Rodobens', 'Santander',
             'Unicoob', 'Embracon', 'GM/Chevrolet', 'Bamaq', 'Banco do Brasil','']

lista_adm_nao_parceira = ['Ancora','CNK','Disal','Honda','Jockey','Multimarcas','Reserva','VolksWagen','Yamaha','Outra','']

def login_user():
    if not usuario.get() in lista_usuarios:
        ent_user = tk.Entry(frm_login, textvariable=usuario, width=40,font=font3,highlightthickness=2,highlightbackground='red').grid(columnspan=2,row=1,column=0)
        lbl_wrong_user = tk.Label(frm_login, text='usuário não cadastrado', font=font3, bg=azul, fg=branco, pady=5).grid(row=0,column=1,sticky='W')
    else:
        ent_user = tk.Entry(frm_login, textvariable=usuario, width=40,font=font3,highlightthickness=2,highlightbackground='gray').grid(columnspan=2,row=1,column=0)
        lbl_wrong_user = tk.Label(frm_login, text='                                               ', font=font3, bg=azul, fg=branco, pady=5).grid(row=0,column=1,sticky='W')
        login_senha()

def login_senha():
    if senha.get() != lista_usuarios[usuario.get()][1]:
        ent_psd = tk.Entry(frm_login, textvariable=senha, width=40, show='*', font=font3,highlightthickness=2,highlightbackground='red').grid(columnspan=2, row=3,column=0)
        lbl_wrong_psd = tk.Label(frm_login, text='senha incorreta',font=font3,bg=azul,fg=branco,pady=5).grid(row=2,column=1,sticky='W')

    else:
        global atendente
        atendente = lista_usuarios[usuario.get()][0]
        abrirtela()

def abrirtela():
    # carregar planilha de registros
    global wb,sheet_todos,sheet_criticos
    wb = load_workbook(filename="assets/Registro de atendimentos.xlsx")
    sheet_todos = wb['Todos']
    sheet_criticos = wb['Críticos']

    global df
    df = pd.DataFrame(sheet_todos.values)

    total_atendimentos_hoje = 0
    atendimentos_usuario_hoje = 0
    resolvidos_hoje_usuario = 0
    acompanhar_hoje_usuario = 0
    criticos_hoje_usuario = 0

    for row in range(0, len(df)):
        if df.loc[row, 1] == str(datetime.now().strftime('%d/%b/%Y')):
            total_atendimentos_hoje += 1
            if df.loc[row, 3] == atendente:
                atendimentos_usuario_hoje += 1
                if df.loc[row, 11] == 'Resolvido':
                    resolvidos_hoje_usuario += 1
                elif df.loc[row, 11] == 'Acompanhar':
                    acompanhar_hoje_usuario += 1
                else:
                    criticos_hoje_usuario += 1


    # limpa a tela e abre área logada
    for widget in win.winfo_children():
        widget.destroy()

    win.title(f'Sistema de Atendimento Consorciei - Home')
    win.geometry('800x600')
    azul_escuro = '#00367E'

    frm_img = tk.Frame(win, padx=100, bg=azul)
    frm_img.pack(expand=TRUE)
    logo = tk.Label(frm_img, image=new_image, bg=azul).pack()

    lbl_ola = tk.Label(win, text=f'Olá, {atendente}!', font=font3, bg=azul, fg=branco).pack(expand=TRUE)

    frm_infos = tk.Frame(win, padx=100, bg=azul_escuro)
    frm_infos.pack(expand=TRUE)

    lbl_2 = tk.Label(frm_infos, text=f'\nTotal de atendimentos hoje: {total_atendimentos_hoje}', font=font3, bg=azul_escuro, fg=branco).pack()
    lbl_3 = tk.Label(frm_infos, text=f'Atendimentos realizados por você: {atendimentos_usuario_hoje}\n\n', font=font3, bg=azul_escuro, fg=branco).pack()

    lbl_4 = tk.Label(frm_infos, text=f'Resolvidos: {resolvidos_hoje_usuario}', font=font3, bg=azul_escuro, fg=branco).pack()
    lbl_5 = tk.Label(frm_infos, text=f'Acompanhar: {acompanhar_hoje_usuario}', font=font3, bg=azul_escuro, fg=branco).pack()
    lbl_6 = tk.Label(frm_infos, text=f'Críticos: {criticos_hoje_usuario}\n', font=font3, bg=azul_escuro, fg=branco).pack()

    botao = tk.Button(win, text='Atender', command=atender, fg='#202124', bg=branco, font=font3).pack(expand=TRUE,pady=20)

def atender():
    win.title('Sistema de Atendimento Consorciei - Registrar')
    win.geometry('800x600')

    # limpa a tela para o atendimento
    for widget in win.winfo_children():
        widget.destroy()

    # criar variaveis de registro
    global datainicio,nome,telefone,cpf,email,adm,adm_np,motivo,status,obs
    datainicio = datetime.now()
    nome, telefone, cpf, email, adm, adm_np, motivo, obs, status = StringVar(),StringVar(),StringVar(),StringVar(),StringVar(),StringVar(),StringVar(),StringVar(),StringVar()


    # campos de registro
    azul_escuro = '#00367E'
    frm_sup = tk.Frame(win,bg=azul_escuro)
    frm_sup.pack(pady=25,ipadx=20,ipady=20)
    frm_geral = tk.Frame(frm_sup,bg=azul_escuro)
    frm_geral.pack(pady=5,ipady=5,ipadx=5)
    lbl_nome = tk.Label(frm_geral, text='Nome:', font=font3, bg=azul_escuro, fg=branco, pady=5).grid(row=0, column=0,sticky='W')
    ent_nome = tk.Entry(frm_geral, textvariable=nome, width=30, font=font3).grid(row=1, column=0,sticky=W)

    lbl_telefone = tk.Label(frm_geral, text='Telefone:', font=font3, bg=azul_escuro, fg=branco, pady=5).grid(padx=20,row=0, column=1,sticky='W')
    ent_telefone = tk.Entry(frm_geral, textvariable=telefone, width=30, font=font3).grid(padx=20,row=1, column=1,sticky=W)

    lbl_cpf = tk.Label(frm_geral, text='CPF:', font=font3, bg=azul_escuro, fg=branco, pady=5).grid(row=2, column=0,sticky='W')
    ent_cpf = tk.Entry(frm_geral, textvariable=cpf, width=30, font=font3).grid(row=3, column=0,sticky=W)

    lbl_email = tk.Label(frm_geral, text='E-mail:', font=font3, bg=azul_escuro, fg=branco, pady=5).grid(padx=20,row=2, column=1,sticky='W')
    ent_email = tk.Entry(frm_geral, textvariable=email, width=30, font=font3).grid(padx=20,row=3, column=1,sticky=W)

    lbl_motivo = tk.Label(frm_geral, text='Motivo:', font=font3, bg=azul_escuro, fg=branco, pady=5).grid(row=4,column=0,sticky=W)
    cb_motivo = ttk.Combobox(frm_geral, values=lista_motivo, font=font3, textvariable=motivo, width=60)
    cb_motivo.set('1. Vender o consorcio/conhecer como trabalhamos e calculamos o preço')
    cb_motivo.grid(columnspan=2, row=5, column=0, pady=5, sticky=W)

    lbl_adm = tk.Label(frm_geral, text='Administradora:', font=font3, bg=azul_escuro, fg=branco, pady=5).grid(row=6, column=0, sticky='W')
    cb_adm = ttk.Combobox(frm_geral,values=lista_adm,font=font3,textvariable=adm,width=28)
    cb_adm.set('Itaú')
    cb_adm.grid(row=7, column=0,pady=5,sticky=W)

    lbl_adm_nao_parceira = tk.Label(frm_geral, text='Administradora (não parceira):', font=font3, bg=azul_escuro, fg=branco, pady=5).grid(padx=20,row=6, column=1,sticky='W')
    cb_adm_np = ttk.Combobox(frm_geral, values=lista_adm_nao_parceira, font=font3, textvariable=adm_np, width=28)
    cb_adm_np.set('')
    cb_adm_np.grid(row=7, column=1, pady=5,padx=20, sticky=W)

    lbl_obs = tk.Label(frm_geral, text='Observações:', font=font3, bg=azul_escuro, fg=branco, pady=5).grid(columnspan=2,row=8,column=0,sticky='W')
    ent_obs = tk.Entry(frm_geral, textvariable=obs, width=62, font=font3).grid(columnspan=2, row=9, column=0, sticky=W)

    lbl_status = tk.Label(frm_geral, text='Status:', font=font3, bg=azul_escuro, fg=branco, pady=5).grid(columnspan=2,row=10, column=0)
    frm_btn = tk.Frame(frm_geral,bg=azul_escuro)
    frm_btn.grid(columnspan=2)
    status1 = tk.Radiobutton(frm_btn,text='Resolvido',variable=status,value='Resolvido',indicatoron=TRUE, font=font3,bg=branco).pack(side=LEFT,padx=10,pady=5,ipadx=5,ipady=2)
    status2 = tk.Radiobutton(frm_btn,text='Acompanhar',variable=status,value='Acompanhar',indicatoron=TRUE, font=font3, bg=branco).pack(side=LEFT,padx=10,pady=5,ipadx=5,ipady=2)
    status3 = tk.Radiobutton(frm_btn,text='Crítico',variable=status,value='Crítico',indicatoron=TRUE, font=font3,bg=branco).pack(side=LEFT,padx=10,pady=5,ipadx=5,ipady=2)
    status.set('Acompanhar')
    botao = tk.Button(win, text='Registrar', command=registrar, fg='#202124', bg=branco, font=font3).pack(side=RIGHT,padx=10,pady=5,ipadx=5,ipady=2)
    botao = tk.Button(win, text='Cancelar', command=abrirtela, fg='#202124', bg=branco, font=font3).pack(side=RIGHT,padx=10,pady=5,ipadx=5,ipady=2)

def registrar():
    try:
        datafim = datetime.now()
        duracao = datafim - datainicio

        xdiasemana = datainicio.strftime('%A')
        xdatainicio = datainicio.strftime("%d/%b/%Y")
        xhora = datainicio.strftime("%H:%M:%S")

        xnome = nome.get()
        xtelefone = telefone.get()
        xcpf = cpf.get()
        xemail = email.get()
        xadm = adm.get()
        if adm_np.get() != '':
            xadm = adm_np.get()
        xmotivo = motivo.get()
        xstatus = status.get()
        xobs = obs.get()

        if len(xtelefone) >= 8 and xstatus != '':
            sheet_todos.append([xdiasemana,xdatainicio,xhora,atendente,duracao,xnome,xtelefone,xcpf,xemail,xadm,xmotivo,xstatus,xobs])
            letras = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L','M']

            if xstatus == 'Acompanhar':
                for letra in letras:
                    sheet_todos[f'{letra}{len(df)+1}'].fill = PatternFill("solid", fgColor="FFDC6D")

            elif xstatus == 'Crítico':
                for letra in letras:
                    sheet_todos[f'{letra}{len(df)+1}'].fill = PatternFill("solid", fgColor="FF8585")
                    sheet_criticos.append([xdiasemana, xdatainicio, xhora, atendente, duracao, xnome, xtelefone, xcpf, xemail, xadm, xmotivo, xstatus, xobs])

            # save the file
            wb.save(filename="assets/Registro de atendimentos.xlsx")
            abrirtela()
        else:
            pyautogui.alert(text='Nenhum registro anotado. Confira os campos', title='Aviso', button='OK')
    except PermissionError:
        pyautogui.alert(text='Feche a planilha de atendimentos para registrar', title='Aviso', button='OK')

win = Tk()
win.geometry('600x400')
win.title('Sistema de Atendimento Consorciei - Login')
azul = '#2863D0'
branco = '#FFF'
win.configure(bg=azul)
font3 = tkFont.Font(family="Arial Black", size=12)
usuario = tk.StringVar()
senha = tk.StringVar()

frm_img = tk.Frame(win,padx=100,bg=azul)
frm_img.pack(expand=TRUE)
frm_login = tk.Frame(win,bg=azul)
frm_login.pack(expand=TRUE)

img= (Image.open("assets/img.png"))
resized_image= img.resize((335,180), Image.ANTIALIAS)
new_image= ImageTk.PhotoImage(resized_image)

logo = tk.Label(frm_img,image=new_image,bg=azul).pack()

lbl_user = tk.Label(frm_login, text='Digite o seu usuário:',font=font3,bg=azul,fg=branco,pady=5).grid(row=0,column=0,sticky='W')
ent_user = tk.Entry(frm_login, textvariable=usuario, width=40,font=font3).grid(columnspan=2,row=1,column=0)
lbl_wrong_user = tk.Label(frm_login, text='',font=font3,bg=azul,fg=branco,pady=5).grid(row=0,column=1,sticky='W')

lbl_psd = tk.Label(frm_login, text='Digite a sua senha:',font=font3,bg=azul,fg=branco,pady=5).grid(row=2,column=0,sticky='W')
ent_psd = tk.Entry(frm_login, textvariable=senha, width=40, show='*',font=font3).grid(columnspan=2,row=3,column=0)
lbl_wrong_psd = tk.Label(frm_login, text='',font=font3,bg=azul,fg=branco,pady=5).grid(row=2,column=1,sticky='W')

botao = tk.Button(frm_login, text='Login', command=login_user,fg='#202124',bg=branco,font=font3).grid(columnspan=2,row=4,column=0,pady=20)

win.mainloop()